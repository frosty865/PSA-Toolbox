#!/usr/bin/env python3
"""
CORPUS: Ingest VOFC Library XLSX

Ingests VOFC_Library.xlsx into CORPUS with strict traceability.
Creates one document per sheet with sheet/row locators.

HARD RULE: Only writes to CORPUS database (yylslokiaovdythzrbgt)
"""

import os
import sys
import hashlib
import argparse
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import re

import psycopg2
from urllib.parse import urlparse
import openpyxl
from openpyxl.utils import get_column_letter

# Add parent directory to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

from tools.corpus.source_set import (
    get_corpus_db_connection,
    require_active_source_set,
    ALLOWED_SOURCE_SETS
)

def load_env_file(filepath: str):
    """Load environment variables from .env.local file."""
    if not os.path.exists(filepath):
        return
    
    with open(filepath, 'r') as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith('#') and '=' in line:
                key, value = line.split('=', 1)
                os.environ[key.strip()] = value.strip().strip('"').strip("'")

def sha256_hash(data: str) -> str:
    """Calculate SHA256 hash of string."""
    return hashlib.sha256(data.encode('utf-8')).hexdigest()

def normalize_header(header: str) -> str:
    """Normalize header text for matching."""
    return re.sub(r'[^\w\s]', '', header.lower().strip())

def find_column_indices(sheet, header_row: int) -> Dict[str, int]:
    """
    Find column indices for key fields.
    
    Returns dict mapping normalized field names to column indices (0-based).
    """
    indices = {}
    
    # Expected fields (normalized)
    target_fields = {
        'category': ['category', 'parent question', 'question'],
        'vulnerability': ['vulnerability', 'child question', 'answer'],
        'ofc': ['options for consideration', 'option', 'ofc', 'recommendation']
    }
    
    # Read header row
    for col_idx, cell in enumerate(sheet[header_row], start=1):
        if cell.value:
            header_text = normalize_header(str(cell.value))
            
            # Match against target fields
            for field_name, patterns in target_fields.items():
                if any(pattern in header_text for pattern in patterns):
                    if field_name not in indices:
                        indices[field_name] = col_idx - 1  # Convert to 0-based
                    break
    
    return indices

def extract_rows_from_sheet(sheet, header_row: int = 1) -> List[Dict]:
    """
    Extract rows from sheet with field mapping.
    
    Returns list of dicts with:
    - category
    - vulnerability
    - ofc
    - row_number (1-based Excel row)
    - raw_row_data (all cell values)
    """
    col_indices = find_column_indices(sheet, header_row)
    
    if not col_indices:
        return []
    
    rows = []
    
    # Iterate data rows (skip header)
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        row_data = {}
        raw_values = []
        
        # Extract field values
        for field_name, col_idx in col_indices.items():
            cell = sheet.cell(row=row_idx, column=col_idx + 1)
            value = str(cell.value).strip() if cell.value else ''
            row_data[field_name] = value
            raw_values.append(value)
        
        # Skip if all fields are empty
        if not any(row_data.values()):
            continue
        
        # Skip if vulnerability or OFC is blank (with warning)
        if not row_data.get('vulnerability') or not row_data.get('ofc'):
            print(
                f"⚠️  Warning: Skipping row {row_idx} in sheet '{sheet.title}': "
                f"missing vulnerability or OFC",
                file=sys.stderr
            )
            continue
        
        row_data['row_number'] = row_idx
        row_data['raw_row_data'] = raw_values
        
        rows.append(row_data)
    
    return rows

def create_document_for_sheet(
    conn,
    sheet_name: str,
    source_id: str,
    rows: List[Dict],
    xlsx_path: str
) -> Tuple[str, int]:
    """
    Create one document per sheet.
    
    Returns: (document_id, chunk_count)
    """
    cur = conn.cursor()
    
    try:
        # Build document title
        title = f"VOFC Library — {sheet_name}"
        
        # Build content hash from sheet name + all row data
        content_str = sheet_name + '\n' + '\n'.join(
            '|'.join(str(v) for v in row['raw_row_data'])
            for row in rows
        )
        content_hash = sha256_hash(content_str)
        
        # Check if document already exists (idempotency)
        cur.execute("""
            SELECT document_id 
            FROM public.documents 
            WHERE source_set = 'VOFC_LIBRARY'
                AND title = %s
                AND file_hash = %s
        """, (title, content_hash))
        
        existing = cur.fetchone()
        if existing:
            document_id = existing[0]
            print(f"  Document already exists: {document_id}")
            
            # Count existing chunks
            cur.execute("""
                SELECT COUNT(*) FROM public.document_chunks
                WHERE document_id = %s
            """, (document_id,))
            chunk_count = cur.fetchone()[0]
            
            return str(document_id), chunk_count
        
        # Create new document
        cur.execute("""
            INSERT INTO public.documents (
                source_id, title, file_path, file_hash,
                source_set, ingested_at
            )
            VALUES (%s, %s, %s, %s, 'VOFC_LIBRARY', now())
            RETURNING document_id
        """, (source_id, title, xlsx_path, content_hash))
        
        document_id = str(cur.fetchone()[0])
        
        # Create chunks (one per row)
        chunk_count = 0
        for row in rows:
            # Build chunk text
            chunk_text = f"Category: {row.get('category', '')}\n"
            chunk_text += f"Vulnerability: {row.get('vulnerability', '')}\n"
            chunk_text += f"OFC: {row.get('ofc', '')}"
            
            # Chunk hash
            chunk_hash = sha256_hash(chunk_text)
            
            # Check if chunk exists (idempotency by document_id + chunk_index)
            cur.execute("""
                SELECT chunk_id FROM public.document_chunks
                WHERE document_id = %s AND chunk_index = %s
            """, (document_id, row['row_number']))
            
            if cur.fetchone():
                continue  # Skip existing chunk
            
            # Build locator: "sheet=<sheet_name>;row=<excel_row_number>"
            locator = f"sheet={sheet_name};row={row['row_number']}"
            
            # Insert chunk (chunk_index = Excel row number for traceability)
            # Note: page_number is NULL for XLSX (not applicable)
            cur.execute("""
                INSERT INTO public.document_chunks (
                    document_id, chunk_index, chunk_text,
                    source_set, locator_type, locator, created_at
                )
                VALUES (%s, %s, %s, 'VOFC_LIBRARY', 'XLSX', %s, now())
            """, (document_id, row['row_number'], chunk_text, locator))
            
            chunk_count += 1
        
        conn.commit()
        
        return document_id, chunk_count
        
    finally:
        cur.close()

def ingest_xlsx(xlsx_path: str, dry_run: bool = False) -> Dict:
    """
    Ingest VOFC_Library.xlsx into CORPUS.
    
    Returns summary dict.
    """
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"XLSX file not found: {xlsx_path}")
    
    conn = get_corpus_db_connection()
    
    try:
        # Enforce active source set
        active_set = require_active_source_set(conn, expected='VOFC_LIBRARY')
        
        # Load workbook
        print(f"Loading workbook: {xlsx_path}")
        workbook = openpyxl.load_workbook(xlsx_path, data_only=True)
        
        # Get or create canonical source
        cur = conn.cursor()
        cur.execute("""
            SELECT source_id FROM public.canonical_sources
            WHERE title = 'VOFC Library' AND source_type = 'GUIDE'
            LIMIT 1
        """)
        
        source_row = cur.fetchone()
        if source_row:
            source_id = source_row[0]
        else:
            cur.execute("""
                INSERT INTO public.canonical_sources (
                    title, source_type, citation_text
                )
                VALUES ('VOFC Library', 'GUIDE', 'VOFC Library (XLSX)')
                RETURNING source_id
            """)
            source_id = cur.fetchone()[0]
            conn.commit()
        
        cur.close()
        
        # Process each sheet
        documents_created = 0
        documents_existing = 0
        chunks_created = 0
        sheets_processed = 0
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            if sheet.max_row <= 1:
                print(f"  Skipping empty sheet: {sheet_name}")
                continue
            
            print(f"  Processing sheet: {sheet_name}")
            
            # Extract rows
            rows = extract_rows_from_sheet(sheet, header_row=1)
            
            if not rows:
                print(f"    No valid rows found")
                continue
            
            if dry_run:
                print(f"    [DRY RUN] Would create document with {len(rows)} chunks")
                chunks_created += len(rows)
                sheets_processed += 1
                continue
            
            # Create document for sheet
            document_id, chunk_count = create_document_for_sheet(
                conn, sheet_name, source_id, rows, xlsx_path
            )
            
            if chunk_count > 0:
                documents_created += 1
                chunks_created += chunk_count
                print(f"    Created document {document_id}: {chunk_count} chunks")
            else:
                documents_existing += 1
                print(f"    Document exists: {document_id}")
            
            sheets_processed += 1
        
        workbook.close()
        
        return {
            'xlsx_path': xlsx_path,
            'sheets_processed': sheets_processed,
            'documents_created': documents_created,
            'documents_existing': documents_existing,
            'chunks_created': chunks_created,
            'active_source_set': active_set,
            'dry_run': dry_run
        }
        
    finally:
        conn.close()

def main():
    parser = argparse.ArgumentParser(description='Ingest VOFC Library XLSX into CORPUS')
    parser.add_argument('--xlsx', required=True, help='Path to VOFC_Library.xlsx')
    parser.add_argument('--dry-run', action='store_true', help='Print counts without DB writes')
    
    args = parser.parse_args()
    
    try:
        result = ingest_xlsx(args.xlsx, dry_run=args.dry_run)
        
        print(f"\n✅ Ingestion complete:")
        print(f"   Sheets processed: {result['sheets_processed']}")
        print(f"   Documents created: {result['documents_created']}")
        print(f"   Documents existing: {result['documents_existing']}")
        print(f"   Chunks created: {result['chunks_created']}")
        print(f"   Active source set: {result['active_source_set']}")
        
        if result['dry_run']:
            print(f"\n   [DRY RUN] No database writes performed")
        
    except Exception as e:
        print(f"❌ Error: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == '__main__':
    main()


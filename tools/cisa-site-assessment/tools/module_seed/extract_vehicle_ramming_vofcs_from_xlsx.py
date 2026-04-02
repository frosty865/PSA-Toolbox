#!/usr/bin/env python3
"""
Extract vehicle ramming VOFCs from XLSX file.

IMPORTANT: This script must be run in a virtual environment (venv).
  Activate venv first:
    Windows: .venv\Scripts\activate
    Linux/Mac: source .venv/bin/activate
"""
import os, json, re, hashlib
from pathlib import Path
import sys

# Check if running in a virtual environment
if not hasattr(sys, 'real_prefix') and not (hasattr(sys, 'base_prefix') and sys.base_prefix != sys.prefix):
    print("ERROR: This script must be run in a virtual environment (venv)")
    print("\nTo activate venv:")
    print("  Windows: .venv\\Scripts\\activate")
    print("  Linux/Mac: source .venv/bin/activate")
    print("\nOr create a venv if it doesn't exist:")
    print("  python -m venv .venv")
    sys.exit(1)

import openpyxl

MODULE_CODE = "MODULE_VEHICLE_RAMMING_SAT"
OUT_JSON = Path("tools/module_seed/MODULE_VEHICLE_RAMMING_SAT_vofcs.json")

KEYWORDS = [
  r"\bvehicle ramming\b", r"\bramming attack\b", r"\bhostile vehicle\b", r"\bhvm\b",
  r"\bcrash[- ]rated\b", r"\bbollard\b", r"\bvehicle barrier\b", r"\banti-ram\b",
  r"\bstandoff\b", r"\bplanter\b", r"\bjersey barrier\b"
]

def sha256_file(p: Path) -> str:
  h = hashlib.sha256()
  h.update(p.read_bytes())
  return h.hexdigest()

def norm(s: str) -> str:
  return re.sub(r"\s+", " ", (s or "").strip())

def matches(text: str) -> bool:
  t = (text or "").lower()
  for pat in KEYWORDS:
    if re.search(pat, t, flags=re.IGNORECASE):
      return True
  return False

def main():
  xlsx_path = os.environ.get("VOFC_LIBRARY_XLSX")
  if not xlsx_path:
    raise SystemExit("VOFC_LIBRARY_XLSX env var not set")

  xlsx = Path(xlsx_path)
  if not xlsx.exists():
    raise SystemExit(f"VOFC library not found: {xlsx}")

  wb = openpyxl.load_workbook(xlsx, data_only=True)
  xlsx_sha = sha256_file(xlsx)

  items = []
  seen = set()

  for sheet in wb.worksheets:
    for r in range(1, sheet.max_row + 1):
      vals = []
      for c in range(1, sheet.max_column + 1):
        v = sheet.cell(row=r, column=c).value
        if v is None:
          continue
        s = norm(str(v))
        if s:
          vals.append(s)
      if not vals:
        continue

      row_text = " | ".join(vals)
      if not matches(row_text):
        continue

      title = vals[0][:140]
      vofc_text = max(vals, key=len)

      key = norm(vofc_text).lower()
      if key in seen:
        continue
      seen.add(key)

      items.append({
        "title": title,
        "vofc_text": vofc_text,
        "tags": ["vehicle-ramming", "hvm"],
        "citation": {
          "locator_type": "XLSX_SHEET_ROW",
          "locator_json": {"sheet": sheet.title, "row": r}
        }
      })

  OUT_JSON.parent.mkdir(parents=True, exist_ok=True)
  OUT_JSON.write_text(json.dumps({
    "module_code": MODULE_CODE,
    "source": {
      "source_path": str(xlsx),
      "source_sha256": xlsx_sha,
      "title": xlsx.name
    },
    "items": items
  }, indent=2), encoding="utf-8")

  print(f"✓ Extracted {len(items)} module VOFC candidate(s) -> {OUT_JSON}")

if __name__ == "__main__":
    main()
